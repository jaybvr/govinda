from django.shortcuts import render,	redirect
from django.http import HttpResponse
from	.models	import *
from	django.template	import	loader
from	django.shortcuts	import	render,	get_object_or_404
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login,	logout
from django.utils import timezone
from	django.core.files.storage	import	default_storage
from django.core.mail import send_mail
from	django.contrib	import	messages
#from celery.decorators import task
#from .tasks import *
#from celery.result import AsyncResult
from requests.packages.urllib3.exceptions import InsecureRequestWarning
from urllib3.exceptions import InsecureRequestWarning

import	string
import	datetime
import	csv
import	os
import	subprocess
import	openpyxl
import	functools
import	pdb
import logging
import	re
import requests
import	json
import time

logger = logging.getLogger('django')

def pingStatus(hip):
	status = ""
	result = ""
	status , result = subprocess.getstatusoutput("ping -c1 -w2 " + str(hip))
	if status == 0:
		return True
	else:
		return False

def index(request):
							return render(request,'hwt/index.html')

def list_main(request):
		hosts=""
		hmcs=""
		storages=""
		fabrics=""
		networks=""
		hosts=Hosts.objects.all()
		hmcs=Hmcs.objects.all()
		storages=Storages.objects.all()
		fabrics=Fabrics.objects.all()
		networks=Networks.objects.all()
		return render(request,"hwt/main.html", {'hcount':len(hosts),'hmcount':len(hmcs),'scount':len(storages),'fcount':len(fabrics),'ncount':len(networks)})

def	login_user(request):
							error=""
							if	request.POST:
										uid=request.POST['uid']
										pwd	=request.POST['pwd']
										user1 = authenticate(username=uid, password=pwd)
										if	user1	is	not	None:
																login(request,user1)
																hosts=""
																hmcs=""
																storages=""
																fabrics=""
																networks=""
																hosts=Hosts.objects.all()
																hmcs=Hmcs.objects.all()
																storages=Storages.objects.all()
																fabrics=Fabrics.objects.all()
																networks=Networks.objects.all()
																return render(request,"hwt/main.html", {'hcount':len(hosts),'hmcount':len(hmcs),'scount':len(storages),'fcount':len(fabrics),'ncount':len(networks)})
										else:
														error="Invalid	Credentials"
							return	render(request,'hwt/Login.html',{'error':error})

def	logout_user(request):
				logout(request)
				success="Successfully"
				return	render(request,'hwt/Login.html',{'success':success})

@login_required
def	export_host_list(request):
					response=HttpResponse(content_type='text/csv')
					response['Content-Disposition']	=	'attachment;	filename=PowerVCHosts'	+	str(datetime.datetime.now())	+	'.csv'
					writer=csv.writer(response)
					writer.writerow(['FSP_/_BMC_IP','Serial','Location','CEC_Name','FSP_Credentials',	'Host_Type',	'Firmware_Level',	'HMC_IP',	'Novalink_IP',	'VIOS1',	'VIOS2',	'Proc_/_Memory',	'Model',	'Network_Ports',	'Fabric_Switch1',	'Fabric_Switch2',	'Primary_VLAN',	'Connected_Storages',	'Owner',	'Squad',	'LAB'])
					hosts=Hosts.objects.all()
					for	h	in	hosts:
								writer.writerow([h.cec_fsp_bmc_ip,h.cec_serial,h.cec_location,h.cec_name,h.cec_fsp_credentials,h.cec_type,h.cec_firmware,h.cec_hmc_ip,h.cec_neo_ip,h.cec_vios1,h.cec_vios2,h.cec_proc_memory,h.cec_model,h.cec_network_ports,h.cec_fabric1,h.cec_fabric2,h.cec_pvlan,h.cec_storages,h.cec_owner,h.cec_squad,h.cec_lab])
					logger.info(" Export Host	List Successful ,	{}".format(request.user.username))
					return	response								
					
def loggs1(message,type):
    if type== "error":
    	logger.error(message)
    elif type == "info":
        logger.info(meessage)

@login_required
def	add_host(request):
					error=""
					host_ip=""
					htypes=[]
					squads=[]
					sq=""
					htypes=HostTypes.objects.values_list('name',flat=True)
					squads=Squads.objects.values_list('name',flat=True)
					if	request.POST:
								host_ip=str(request.POST['cec_fsp_bmc']).strip()
								if	(host_ip):
										try:
													host = Hosts.objects.get(pk=host_ip)
													error="Host	with	this	IP	already	present"
													#loggs.delay("	Host	with	IP	{}	is	already	present,	{}".format(host_ip,request.user.username),"error")
													logger.error("	Host	with	IP	{}	is	already	present,	{}".format(host_ip,request.user.username))
										except	Hosts.DoesNotExist:																																	
																		if	'cec_squad'	not	in	request.POST:

																			
																								sq="Pool"
																		else:
																							sq=	request.POST['cec_squad']
																		host=Hosts(cec_name=request.POST['cec_name'],
																		cec_ip_ping= pingStatus(host_ip),
																		cec_serial=request.POST['cec_serial'],
																		cec_location=request.POST['cec_location'],
																		cec_fsp_bmc_ip=host_ip,
																		cec_fsp_credentials=request.POST['cec_fsp_bmc_uid'],
																		cec_type=request.POST['cec_type'],
																		cec_firmware=request.POST['cec_firmware'],
																		cec_hmc_ip=request.POST['cec_hmc'],
																		cec_neo_ip=request.POST['cec_novalink'],
																		cec_vios1=request.POST['cec_vios1'],
																		cec_vios2=request.POST['cec_vios2'],
																		cec_proc_memory=request.POST['cec_proc_mem'],
																		cec_model=request.POST['cec_model'],
																		cec_network_ports=request.POST['cec_nw_ports'],
																		cec_fabric1=request.POST['cec_fabric1'],
																		cec_fabric2=request.POST['cec_fabric2'],
																		cec_pvlan=request.POST['cec_pvlan'],
																		cec_storages=request.POST['cec_storages'],
																		cec_owner=request.POST['cec_owner'],
																		cec_squad=sq,
																		cec_lab=request.POST['cec_lab'])
																		host.save()
																		hname=str(request.POST['cec_hmc'])
																		logger.info("	Host	with	IP	{}	is	added,	{}".format(host_ip,request.user.username))
																		return	redirect('list_hosts')
								else:
											error="Mandatory	Fields	are	missing"
					return render(request,'hwt/AddHost.html',{'error':error,'htypes':htypes,'squads':squads})


@login_required
def	edit_host(request):
						error=""
						htypes=[]
						squads=[]
						sq=""
						htypes=HostTypes.objects.values_list('name',flat=True)
						squads=Squads.objects.values_list('name',flat=True)
						
						if	request.POST:
									if	'edit_id'	in	str(request.POST):
												try:
																ehost = Hosts.objects.get(pk=str(request.POST['edit_id']).strip())
																return render(request,'hwt/EditHost.html',{'ehost':ehost,'htypes':htypes,'squads':squads})
												except	Hosts.DoesNotExist:																																	
																error="Host	with	this	IP	Doesn't	exist"
																logger.error("Edit Host	with IP	{}	is failed as host doesn't exist,	{}".format(str(request.POST['edit_id']),request.user.username))

									elif	'cec_fsp_bmc_ip'	in	str(request.POST):
												if	'cec_squad'	not	in	request.POST:
																								sq="Pool"
												else:
																					sq=	request.POST['cec_squad']
																
												host=get_object_or_404(Hosts,	pk=str(request.POST['cec_fsp_bmc_ip']).strip())
												host.cec_serial=request.POST['cec_serial']
												host.cec_location=request.POST['cec_location']
												host.cec_name=request.POST['cec_name']
												host.cec_fsp_credentials=request.POST['cec_fsp_bmc_uid']
												host.cec_type=request.POST['cec_type']
												host.cec_firmware=request.POST['cec_firmware']
												host.cec_hmc_ip=request.POST['cec_hmc']
												host.cec_neo_ip=request.POST['cec_novalink']
												host.cec_vios1=request.POST['cec_vios1']
												host.cec_vios2=request.POST['cec_vios2']
												host.cec_proc_memory=request.POST['cec_proc_mem']
												host.cec_model=request.POST['cec_model']
												host.cec_network_ports=request.POST['cec_nw_ports']
												host.cec_fabric1=request.POST['cec_fabric1']
												host.cec_fabric2=request.POST['cec_fabric2']
												host.cec_pvlan=request.POST['cec_pvlan']
												host.cec_storages=request.POST['cec_storages']
												host.cec_owner=request.POST['cec_owner']
												host.cec_squad=sq
												host.cec_lab=request.POST['cec_lab']
												host.save()
												#return render(request,'hwt/ListAllHosts.html',{'hosts':Hosts.objects.all()})
												logger.info("Edit Host	with IP	{}	is Successful,	{}".format(str(request.POST['cec_fsp_bmc_ip']),request.user.username))
												return	redirect('list_hosts')
						return render(request,'hwt/EditHost.html',{'error':error,'htypes':htypes,'squads':squads})

@login_required
def	delete_host(request):
				error=""
				if	request.POST:
									if	'delete_id'	in	str(request.POST):
												try:
																dhost = Hosts.objects.get(pk=str(request.POST['delete_id']).strip())
																return render(request,'hwt/DeleteHost.html',{'ehost':dhost})
												except	Hosts.DoesNotExist:
																error="Host	with	this	IP	doesn't	exist"
																logger.error("Delete Host with IP {} is Failed as the host doesn't exist,	{}".format(str(request.POST['delete_id']),request.user.username))

									elif	'cec_fsp_bmc_ip'	in	str(request.POST):
												host=get_object_or_404(Hosts,	pk=str(request.POST['cec_fsp_bmc_ip'].strip()))
												host.delete()
												#return render(request,'hwt/ListAllHosts.html',{'hosts':Hosts.objects.all()})
												logger.info("Delete Host with IP {} is Successful,	{}".format(str(request.POST['cec_fsp_bmc_ip']),request.user.username))
												return	redirect('list_hosts')
				return render(request,'hwt/DeleteHost.html',{'error':error})

@login_required
def	export_hmc_list(request):
					response=HttpResponse(content_type='text/csv')
					response['Content-Disposition']	=	'attachment;	filename=PowerVC_HMC_list'	+	str(datetime.datetime.now())	+	'.csv'
					writer=csv.writer(response)
					writer.writerow(['HMC_Public_IP','Serial','Location','HMC_Credentials','HMC_Private_IP','Public_/_Private','IMM_IP','IMM_Credentials','MTMS','Owner','Squad','LAB'])
					hmcs=Hmcs.objects.all()
					for	h	in	hmcs:
								writer.writerow([h.hmc_public_ip,h.hmc_serial,h.hmc_location,h.hmc_credentials,h.hmc_private_ip,h.hmc_public_private,h.hmc_imm_ip,h.hmc_imm_credentials,h.hmc_mtms,h.hmc_owner,h.hmc_squad,h.hmclab])
					logger.info("Export HMC list is Successful,	{}".format(request.user.username))
					return	response
	
@login_required
def	add_hmc(request):
					error=""
					hmc_ip=""
					squads=[]
					sq=""
					squads=Squads.objects.values_list('name',flat=True)
					if	request.POST:
								hmc_ip=str(request.POST['hmc_public_ip']).strip()
								if	(hmc_ip):
														try:
																	hmc = Hmcs.objects.get(pk=hmc_ip)
																	error="Host	with	this	IP	already	present"
																	logger.error("Add HMC with IP {} is Failed as the HMC is already present, {}".format(hmc_ip,request.user.username))
														except	Hmcs.DoesNotExist:							
																		if	'hmc_squad'	not	in	request.POST:
																								sq="Pool"
																		else:
																				sq=	request.POST['hmc_squad']
															
																		hmc=Hmcs(
																			hmc_serial=request.POST['hmc_serial'],
																			hmc_location=request.POST['hmc_location'],
																			hmc_public_ip=hmc_ip,
																			hmc_ip_ping = pingStatus(hmc_ip),
																			hmc_credentials=request.POST['hmc_credentials'],
																			hmc_private_ip=request.POST['hmc_private_ip'],
																			hmc_public_private=request.POST['hmc_public_private'],
																			hmc_imm_ip=request.POST['hmc_imm_ip'],
																			hmc_imm_credentials=request.POST['hmc_imm_credentials'],
																			hmc_mtms=request.POST['hmc_mtms'],
																			hmc_owner=request.POST['hmc_owner'],
																			hmc_squad=sq,
																			hmclab=request.POST['hmc_lab'])
																		hmc.save()
																		logger.info("Add HMC with IP {} is Successful, {}".format(hmc_ip,request.user.username))
																		return render(request,'hwt/ListHMC.html',{'hmcs':Hmcs.objects.all()})
								else:
											error="Mandatory	Fields	are	required"
					return render(request,'hwt/AddHMC.html',{'error':error,'squads':squads})
	

@login_required
def	edit_hmc(request):
					error=""
					hmc_ip=""
					squads=[]
					sq=""
					squads=Squads.objects.values_list('name',flat=True)
					if	request.POST:
									if	'ehmc_id'	in	str(request.POST):
												try:
																hmc_ip=str(request.POST['ehmc_id']).strip()
																ehmc = Hmcs.objects.get(pk=hmc_ip)
																return render(request,'hwt/EditHMC.html',{'ehmc':ehmc,'squads':squads})
												except	Hmcs.DoesNotExist:																																	
																error="HMC	with	this	IP	Doesn't	exist"
																logger.error("Edit HMC with IP {} is Failed as the HMC is not found, {}".format(hmc_ip,request.user.username))

									elif	'hmc_public_ip'	in	str(request.POST):
																if	'hmc_squad'	not	in	request.POST:
																								sq="Pool"
																else:
																					sq=	request.POST['hmc_squad']
																hmc=get_object_or_404(Hmcs,	pk=str(request.POST['hmc_public_ip']).strip())
																hmc.hmc_serial=request.POST['hmc_serial']
																hmc.hmc_location=request.POST['hmc_location']
																hmc.hmc_credentials=request.POST['hmc_credentials']
																hmc.hmc_private_ip=request.POST['hmc_private_ip']
																hmc.hmc_public_private=request.POST['hmc_public_private']
																hmc.hmc_imm_ip=request.POST['hmc_imm_ip']
																hmc.hmc_imm_credentials=request.POST['hmc_imm_credentials']
																hmc.hmc_mtms=request.POST['hmc_mtms']
																hmc.hmc_owner=request.POST['hmc_owner']
																hmc.hmc_squad=sq
																hmc.hmclab=request.POST['hmc_lab']
																hmc.save()
																logger.info("Edit HMC with IP {} is Successful, {}".format(str(request.POST['hmc_public_ip']),request.user.username))

																return render(request,'hwt/ListHMC.html',{'hmcs':Hmcs.objects.all()})
					return render(request,'hwt/EditHMC.html',{'squads':squads})

@login_required
def	delete_hmc(request):
				error=""
				if	request.POST:
									if	'dhmc_id'	in	str(request.POST):
												try:
																hmc_ip=str(request.POST['dhmc_id']).strip()
																dhmc = Hmcs.objects.get(pk=hmc_ip)
																return render(request,'hwt/DeleteHMC.html',{'ehmc':dhmc})
												except	Hmcs.DoesNotExist:																																	
																error="HMC	with	this	IP	Doesn't	exist"
																logger.error("Delete HMC with IP {} is Failed as the IP doesn't exist, {}".format(hmc_ip,request.user.username))

									elif	'hmc_public_ip'	in	str(request.POST):
																hmc=get_object_or_404(Hmcs,	pk=str(request.POST['hmc_public_ip']).strip())
																hmc.delete()
																logger.info("Delete HMC with IP {} is Successful, {}".format(str(request.POST['hmc_public_ip']),request.user.username))

																return render(request,'hwt/ListHMC.html',{'hmcs':Hmcs.objects.all()})
				return render(request,'hwt/DeleteHMC.html',{'error':error})

@login_required
def	export_storage_list(request):
					response=HttpResponse(content_type='text/csv')
					response['Content-Disposition']	=	'attachment;	filename=PowerVC_Storage_list'	+	str(datetime.datetime.now())	+	'.csv'
					writer=csv.writer(response)
					writer.writerow(['Storage_IP','Serial','Location','Storage_Type','Storage_Credentials','Storage_Port','Storage_REST_API_Server','Storage_LDEV_ID_range','Capacity_in_TB','Service_IP','LAB'])
					storages=Storages.objects.all()
					for	s	in	storages:
								writer.writerow([s.storage_ip,s.storage_serial,s.storage_location,storage_type,s.storage_credentials,s.storage_port,s.storage_rest_api_server,s.storage_ldev_id_range,s.storage_capacity,s.storage_service_ip,s.storage_lab])
					logger.info("Export Storage list is Successful, {}".format(request.user.username))
					return	response


@login_required
def	add_storage(request):
					error=""
					if	request.POST:
								s_ip=str(request.POST['storage_ip']).strip()
								if	(s_ip):
													try:
																storage = Storages.objects.get(pk=s_ip)
																error="Storage	already	present"
																logger.error("Add Storage with IP {} is Failed as the IP already present, {}".format(str(s_ip),request.user.username))

													except	Storages.DoesNotExist:
																storage=Storages(storage_type=request.POST['storage_type'],
																									storage_serial=request.POST['storage_serial'],
																									storage_location=request.POST['storage_location'],
																									storage_ip = s_ip,
																									storage_ip_ping = pingStatus(s_ip),
																									storage_credentials =request.POST['storage_credentials'],
																									storage_port =request.POST['storage_port'],
																									storage_rest_api_server =request.POST['storage_rest_api_server'],
																									storage_ldev_id_range =request.POST['storage_ldev_id_range'],
																									storage_capacity =request.POST['storage_capacity'],
																									storage_service_ip =request.POST['storage_serviceip'],
																									storage_lab =request.POST['storage_lab'])
																storage.save()
																logger.info("Add Storage with IP {} is Successful, {}".format(str(s_ip),request.user.username))
																return render(request,'hwt/ListStorages.html',{'storages':Storages.objects.all()})
								else:
												error=	"Mandatory	Fields	are	required"
					return render(request,'hwt/AddStorage.html',{'error':error})

@login_required
def	edit_storage(request):
					error=""
					if	request.POST:
										if'estorage_id'	in	str(request.POST):
													try:
																		estorage=Storages.objects.get(pk=str(request.POST['estorage_id']).strip())
																		return render(request,'hwt/EditStorage.html',{'estorage':estorage})
													except	Storages.DoesNotExist:
																	error=	"Given	Storage	doesn't	exist"	
																	logger.error("Edit Storage with IP {} is Failed as the IP doesn't exist, {}".format(str(request.POST['estorage_id']),request.user.username))
										elif	'storage_ip'	in	str(request.POST):
													storage=get_object_or_404(Storages,	pk=str(request.POST['storage_ip']).strip())
													storage.storage_type=request.POST['storage_type']
													storage.storage_serial=request.POST['storage_serial']
													storage.storage_location=request.POST['storage_location']
													storage.storage_credentials =request.POST['storage_credentials']
													storage.storage_port =request.POST['storage_port']
													storage.storage_rest_api_server =request.POST['storage_rest_api_server']
													storage.storage_ldev_id_range =request.POST['storage_ldev_id_range']
													storage.storage_capacity =request.POST['storage_capacity']
													storage.storage_service_ip =request.POST['storage_serviceip']
													storage.storage_lab =request.POST['storage_lab']
													storage.save()
													logger.info("Edit Storage with IP {} is Successful, {}".format(str(request.POST['storage_ip']),request.user.username))
													return render(request,'hwt/ListStorages.html',{'storages':Storages.objects.all()})
					return render(request,'hwt/EditStorage.html',{'error':error})

@login_required
def	delete_storage(request):
				error=""
				if	request.POST:
										if'dstorage_id'	in	str(request.POST):
													try:
																		dstorage=Storages.objects.get(pk=str(request.POST['dstorage_id']).strip())
																		return render(request,'hwt/DeleteStorage.html',{'estorage':dstorage})
													except	Storages.DoesNotExist:
																	error=	"Given	Storage	doesn't	exist"
																	logger.error("Delete Storage with IP {} is Failed as the storage doesn't exist, {}".format(str(request.POST['dstorage_id']),request.user.username))

										elif	'storage_ip'	in	str(request.POST):
													storage=get_object_or_404(Storages,	pk=str(request.POST['storage_ip']).strip())
													storage.delete()
													logger.info("Delete Storage with IP {} is Successful, {}".format(str(request.POST['storage_ip'])),request.user.username)
													return render(request,'hwt/ListStorages.html',{'storages':Storages.objects.all()})
				return render(request,'hwt/DeleteStorage.html',{'error':error})

@login_required
def	export_fabric_list(request):
					response=HttpResponse(content_type='text/csv')
					response['Content-Disposition']	=	'attachment;	filename=PowerVC_Fabrics'	+	str(datetime.datetime.now())	+	'.csv'
					writer=csv.writer(response)
					writer.writerow(['Fabric_IP','Serial','Location','Credentials','Brocade_/_CISCO','Primary_/_Secondary','Fabric_Port','VSAN_ID','LAB','Connected_Storwises','Connected_DS8K','Connected_EMC_VMAX','Connected_EMC_VNX','Connected_Hitachi','Connected_XIV'])
					fabrics=Fabrics.objects.all()
					for	f	in	fabrics:
								writer.writerow([f.fabric_ip,fabric_serial,f.fabric_location,f.fabric_credentials,f.fabric_type,f.fabric_primary,f.fabric_primary_port,f.fabric_vsan_id,f.fabric_lab,f.fabric_storwise,f.fabric_ds8k,f.fabric_emc_vmax,f.fabric_emc_vnx,f.fabric_hitachi,f.fabric_xiv])

					logger.info("Export Fabric List is Successful, {}".format(request.user.username))
					return	response
					
@login_required
def	add_fabric(request):
							error=""
							if	request.POST:
										fab_ip=str(request.POST['fabric_ip']).strip()
										if	(fab_ip):
														try:
																		fabr=Fabrics.objects.get(pk=fab_ip)
																		error="Fabric	Already	Present"
																		logger.error("Add Fabric with IP {} is Failed as the fabric already present, {}".format(str(fab_ip),request.user.username))
														except	Fabrics.DoesNotExist:
																			fabr=Fabrics(fabric_serial=request.POST['fabric_serial'],
																			fabric_location=request.POST['fabric_location'],
																			fabric_ip=fab_ip,
																			fabric_ip_ping=pingStatus(fab_ip),
																			fabric_credentials=request.POST['fabric_credentials'],
																			fabric_type=request.POST['fabric_type'],
																			fabric_primary=request.POST['fabric_primary'],
																			fabric_primary_port=request.POST['fabric_port'],
																			fabric_vsan_id=request.POST['fabric_vsan_id'],			
																			fabric_lab=request.POST['fabric_lab'],
																			fabric_storwise=request.POST['fabric_storwise'],
																			fabric_ds8k=request.POST['fabric_ds8k'],
																			fabric_emc_vmax=request.POST['fabric_vmax'],
																			fabric_emc_vnx=request.POST['fabric_vnx'],
																			fabric_hitachi=request.POST['fabric_hitachi'],
																			fabric_xiv=request.POST['fabric_xiv'])
																			fabr.save()
																			logger.info("Add Fabric with IP {} is Successful , {}".format(str(fab_ip),request.user.username))

																			return render(request,'hwt/ListFabrics.html',{'fabrics':Fabrics.objects.all()})
										else:
													error="Mandatory	Fields	are	required"
							return render(request,'hwt/AddFabric.html',{'error':error})

@login_required
def	edit_fabric(request):
							error=""
							if	request.POST:
											if	'efabric_ip'	in	str(request.POST):
															fab_ip=str(request.POST['efabric_ip']).strip()
															try:
																		fabr=Fabrics.objects.get(pk=fab_ip)
																		return render(request,'hwt/EditFabric.html',{'efabric':fabr})
															except	Fabrics.DoesNotExist:
																		error="Fabric	Deosn't	exist"
																		logger.error("Edit Fabric with IP {} is Failed as the Fabric doesn't exist, {}".format(str(fab_ip),request.user.username))

											elif	'fabric_ip'	in	str(request.POST):
															fab=get_object_or_404(Fabrics,	pk=str(request.POST['fabric_ip']).strip())
															fab.fabric_serial=request.POST['fabric_serial']
															fab.fabric_location=request.POST['fabric_location']
															fab.fabric_credentials=request.POST['fabric_credentials']
															fab.fabric_type=request.POST['fabric_type']
															fab.fabric_primary=request.POST['fabric_primary']
															fab.fabric_primary_port=request.POST['fabric_port']
															fab.fabric_vsan_id=request.POST['fabric_vsan_id']		
															fab.fabric_lab=request.POST['fabric_lab']
															fab.fabric_storwise=request.POST['fabric_storwise']
															fab.fabric_ds8k=request.POST['fabric_ds8k']
															fab.fabric_emc_vmax=request.POST['fabric_vmax']
															fab.fabric_emc_vnx=request.POST['fabric_vnx']
															fab.fabric_hitachi=request.POST['fabric_hitachi']
															fab.fabric_xiv=request.POST['fabric_xiv']
															fab.save()
															logger.info("Edit Fabric with IP {} is Successful, {}".format(str(request.POST['fabric_ip']),request.user.username))

															return render(request,'hwt/ListFabrics.html',{'fabrics':Fabrics.objects.all()})
							return render(request,'hwt/EditFabric.html',{'error':error})

@login_required
def	delete_fabric(request):
				error=""
				if	request.POST:
											if	'dfabric_ip'	in	str(request.POST):
															fab_ip=str(request.POST['dfabric_ip']).strip()
															try:
																		fabr=Fabrics.objects.get(pk=fab_ip)
																		return render(request,'hwt/DeleteFabric.html',{'efabric':fabr})
															except	Fabrics.DoesNotExist:
																		error="Fabric	Deosn't	exist"
																		logger.error("Delete Fabric with IP {} is Failed as the Fabric doesn't exist, {}".format(fab_ip,request.user.username))

											elif	'fabric_ip'	in	str(request.POST):
															fab=get_object_or_404(Fabrics,	pk=str(request.POST['fabric_ip']).strip())
															fab.delete()
															logger.info("Delete Fabric with IP {} is Successful, {}".format(str(request.POST['fabric_ip']),request.user.username))
															return render(request,'hwt/ListFabrics.html',{'fabrics':Fabrics.objects.all()})
				return render(request,'hwt/DeleteFabric.html',{'error':error})


@login_required
def	add_image(request):
							error=""
							if	request.POST:
															try:
																		imag=Pvc_Images.objects.get(pk=str(request.POST['image_vol_name']).strip())
																		error="Image	already	present"
																		logger.error("Add Image {} is Failed as the Image already present, {}".format(str(request.POST['image_vol_name']),request.user.username))

															except	Pvc_Images.DoesNotExist:
																		image=Pvc_Images(owner = request.POST['image_owner'],
																									storage = request.POST['image_storage'],
																									operating_system = request.POST['image_os'],
																									image_volume = str(request.POST['image_vol_name']).strip())
																		image.save()
																		logger.info("Add Image {} is Successful, {}".format(str(request.POST['image_vol_name']),request.user.username))
																		return render(request,'hwt/ListImages.html',{'images':Pvc_Images.objects.all()})
							return render(request,'hwt/AddImage.html',{'error':error})

@login_required
def	edit_image(request):
				error=""
				if	request.POST:
											if	'eimage_vol'	in	str(request.POST):
															try:
																		eimage=Pvc_Images.objects.get(pk=str(request.POST['eimage_vol']).strip())
																		return render(request,'hwt/EditImage.html',{'eimage':eimage})
															except	Pvc_Images.DoesNotExist:
																			error="Given	Image	doesn't	exist"
																			logger.error("Edit Image {} is Failed as the Image doesn't exist, {}".format(str(request.POST['eimage_vol']),request.user.username))

											elif	'image_vol'	in	str(request.POST):
															eimag=get_object_or_404(Pvc_Images,	pk=str(request.POST['image_vol']).strip())
															eimag.owner=request.POST['image_owner']
															eimag.storage=request.POST['image_storage']
															eimag.operating_system=request.POST['image_os']
															eimag.save()
															logger.info("Edit Image {} is Successful, {}".format(str(request.POST['image_vol']),request.user.username))
															return render(request,'hwt/ListImages.html',{'images':Pvc_Images.objects.all()})
				return render(request,'hwt/EditImage.html',{'error':error})


@login_required
def	delete_image(request):
				error=""
				if	request.POST:
											if	'dimage_vol'	in	str(request.POST):
															try:
																		dimage=Pvc_Images.objects.get(pk=str(request.POST['dimage_vol']).strip())
																		return render(request,'hwt/DeleteImage.html',{'dimage':dimage})
															except	Pvc_Images.DoesNotExist:
																			error="Given	Image	doesn't	exist"
																			logger.error("Delete Image {} is Failed as the Image doesn't exist, {}".format(str(request.POST['dimage_vol']),request.user.username))

											elif	'image_vol'	in	str(request.POST):
															dimag=get_object_or_404(Pvc_Images,	pk=str(request.POST['image_vol']).strip())
															dimag.delete()
															logger.info("Delete Image {} is Successful, {}".format(str(request.POST['image_vol']),request.user.username))
															return render(request,'hwt/ListImages.html',{'images':Pvc_Images.objects.all()})
				return render(request,'hwt/DeleteImage.html',{'error':error})


@login_required
def	add_network(request):
							error=""
							if	request.POST:
										if	'nw_vlan_id'	in	str(request.POST):
													if	(str(request.POST['nw_vlan_id']).strip()):
																	try:
																					nw=Networks.objects.get(pk=str(request.POST['nw_vlan_id']).strip())
																					error	=	"Network	already	exists"
																					logger.error("Add Network {} is Failed as the Network already present, {}".format(str(request.POST['nw_vlan_id']),request.user.username))

																	except	Networks.DoesNotExist:
																						nw=Networks(vlan_id =request.POST['nw_vlan_id'],
																						subnet_mask =request.POST['nw_subnet_mask'],
																						gateway =request.POST['nw_gateway'],
																						dns1 =request.POST['nw_dns1'],
																						dns2 =request.POST['nw_dns2'])
																						nw.save()
																						logger.info("Add Network {} is Successful, {}".format(str(request.POST['nw_vlan_id']),request.user.username))

																						return render(request,'hwt/ListNetworks.html',{'networks':Networks.objects.all(),'network_nodes':NetworkNodes.objects.all()})
													else:
															error="Mandatory	Fields	are	required"
										elif	'nw_node_ip'	in	str(request.POST):
														node_ip=str(request.POST['nw_node_ip']).strip()
														if	(node_ip):
																	try:
																					nnode=NetworkNodes.objects.get(pk= node_ip)
																					error	=	"Network	Node	already	exists"
																					logger.error("Add Network Node {} is Failed as the Network Node already exists, {}".format(str(request.POST['nw_node_ip']),request.user.username))

																	except	NetworkNodes.DoesNotExist:
																				nnode=NetworkNodes(network_node_ip = node_ip,
																				network_node_ip_ping=pingStatus(node_ip),
																				credentials =request.POST['nw_node_credentials'],
																				x_ppc =request.POST['nw_node_type'],
																				operating_system =request.POST['nw_node_os'])
																				nnode.save()
																				logger.info("Add Network Node {} is Successful, {}".format(str(request.POST['nw_node_ip']),request.user.username))
																				return render(request,'hwt/ListNetworks.html',{'networks':Networks.objects.all(),'network_nodes':NetworkNodes.objects.all()})
														else:
																	error="Mandatory	Fields	are	required"
							return render(request,'hwt/AddNetwork.html',{'error':error})

@login_required
def	edit_network(request):
							error=""
							if	request.POST:
										if	'editvlan_id'	in	str(request.POST):
													if	(str(request.POST['editvlan_id']).strip()):
																	try:
																					nw=Networks.objects.get(pk=str(request.POST['editvlan_id']).strip())
																					return render(request,'hwt/EditNetwork.html',{'nw':nw})
																	except	Networks.DoesNotExist:
																					error	=	"Network	Doesn't	exist"
																					logger.error("Edit Network {} is Failed as the Network doesn't exist, {}".format(str(request.POST['editvlan_id']),request.user.username))

										elif	'evlan_id'	in	str(request.POST):
																	nw=get_object_or_404(Networks,	pk=str(request.POST['evlan_id']).strip())
																	nw.subnet_mask =request.POST['nw_subnet_mask']
																	nw.gateway =request.POST['nw_gateway']
																	nw.dns1 =request.POST['nw_dns1']
																	nw.dns2 =request.POST['nw_dns2']
																	nw.save()
																	logger.info("Edit Network {} is Successful, {}".format(str(request.POST['evlan_id']),request.user.username))

																	return render(request,'hwt/ListNetworks.html',{'networks':Networks.objects.all(),'network_nodes':NetworkNodes.objects.all()})																				
							return render(request,'hwt/EditNetwork.html',{'error':error})

@login_required
def	edit_network_node(request):
				error=""
				if	request.POST:
								if	'enw_node_ip'	in	str(request.POST):
														if	(str(request.POST['enw_node_ip']).strip()):
																				try:
																							nwn=NetworkNodes.objects.get(pk=str(request.POST['enw_node_ip']).strip())
																							return render(request,'hwt/EditNetwork.html',{'nwn':nwn})
																				except	NetworkNodes.DoesNotExist:
																							error	=	"Network	Node	doesn't	exist"
																							logger.error("Edit Network Node {} is Failed as the Network Node doesn't exists, {}".format(str(request.POST['enw_node_ip']),request.user.username))

								elif	'ednw_node_ip'	in	str(request.POST):
																				nwn=get_object_or_404(NetworkNodes,	pk=str(request.POST['ednw_node_ip']).strip())
																				nwn.credentials =request.POST['nw_node_credentials']
																				nwn.x_ppc =request.POST['nw_node_type']
																				nwn.operating_system =request.POST['nw_node_os']
																				nwn.save()
																				logger.info("Edit Network Node {} is Successful, {}".format(str(request.POST['ednw_node_ip']),request.user.username))
																				return render(request,'hwt/ListNetworks.html',{'networks':Networks.objects.all(),'network_nodes':NetworkNodes.objects.all()})
				return render(request,'hwt/EditNetwork.html',{'error':error})



@login_required
def	delete_network(request):
				error=""
				if	request.POST:
										if	'dvlan_id'	in	str(request.POST):
													if	(str(request.POST['dvlan_id']).strip()):
																	try:
																					nw=Networks.objects.get(pk=str(request.POST['dvlan_id']).strip())
																					return render(request,'hwt/DeleteNetwork.html',{'nw':nw})
																	except	Networks.DoesNotExist:
																					error	=	"Network	Doesn't	exist"
																					logger.error("Delete Network  {} is Failed as the Network doesn't exists, {}".format(str(request.POST['dvlan_id']),request.user.username))

													else:
																		error	=	"Mandatory	Fields	are	required"
										elif	'delvlan_id'	in	str(request.POST):
																if	(str(request.POST['delvlan_id']).strip()):
																		nwn=get_object_or_404(Networks,	pk=str(request.POST['delvlan_id']).strip())
																		nwn.delete()
																		logger.info("Delete Network  {} is Successful, {}".format(str(request.POST['delvlan_id']),request.user.username))

																		return render(request,'hwt/ListNetworks.html',{'networks':Networks.objects.all(),'network_nodes':NetworkNodes.objects.all()})
				return render(request,'hwt/DeleteNetwork.html',{'error':error})
				


@login_required
def	delete_network_node(request):
				error=""
				if	request.POST:
										if	'dnw_node_ip'	in	str(request.POST):
													if	(str(request.POST['dnw_node_ip']).strip()):
																			try:
																							nwn=NetworkNodes.objects.get(pk=str(request.POST['dnw_node_ip']).strip())
																							return render(request,'hwt/DeleteNetwork.html',{'nwn':nwn})
																			except	NetworkNodes.DoesNotExist:
																							error	=	"Network	Node	doesn't	exist"
																							logger.error("Delete Network Node {} is Failed as the Network Node doesn't exists, {}".format(str(request.POST['dnw_node_ip']),request.user.username))

													else:
																		error	=	"Mandatory	Fields	are	required"
										elif	'delnw_node_ip'	in	str(request.POST):
																if	(str(request.POST['delnw_node_ip']).strip()):
																		nwn=get_object_or_404(NetworkNodes,	pk=str(request.POST['delnw_node_ip']).strip())
																		nwn.delete()
																		logger.info("Delete Network Node {} is Successful, {}".format(str(request.POST['delnw_node_ip']),request.user.username))

																		return render(request,'hwt/ListNetworks.html',{'networks':Networks.objects.all(),'network_nodes':NetworkNodes.objects.all()})
				return render(request,'hwt/DeleteNetwork.html',{'error':error})

@login_required
def	add_supported_level(request):
						error=""
						if	request.POST:
									if	'release'	in	str(request.POST):
													if	(str(request.POST['release']).strip()):
																	try:
																							re=SupportedLevels.objects.get(pk=str(request.POST['release']).strip())
																							error="Release	already	exists"	
																							logger.error("Add Release {} is Failed as the Release already exists, {}".format(str(request.POST['release']),request.user.username))
																	except	SupportedLevels.DoesNotExist:
																								re=SupportedLevels(pvc_release=str(request.POST['release']).strip(),
																											sld=str(request.POST['sld']),
																											test_plan=str(request.POST['test_plan']),
																											management_node=str(request.POST['management_node']),
																											hmc_level=str(request.POST['hmc_level']),
																											vios_level=str(request.POST['vios_level']),
																											novalink_level=str(request.POST['novalink_level']),
																											novalink_os=str(request.POST['novalink_os']),
																											guest_os=str(request.POST['guest_os']),
																											storages=str(request.POST['storages']),
																											comments=str(request.POST['comments']))
																								re.save()
																								logger.info("Add Release {} is Successful, {}".format(str(request.POST['release']),request.user.username))
																								return render(request,'hwt/ListSupportedLevels.html',{'supp_level':reversed(SupportedLevels.objects.all())})
													else:
															error="Mandatory	Fields	are	required"
						return render(request,'hwt/AddSupportedLevels.html',{'error':error})

@login_required
def	edit_supported_level(request):
				error=""
				if	request.POST:
											if	'erelease_id'	in	str(request.POST):
															try:
																							re=SupportedLevels.objects.get(pk=str(request.POST['erelease_id']).strip())
																							return render(request,'hwt/EditSupportedLevels.html',{'erelease':re})
															except	SupportedLevels.DoesNotExist:
																					error="Given	Release	Doesn't	exist"	
																					logger.error("Edit Release {} is Failed as the Release doesn't exists, {}".format(str(request.POST['erelease_id']),request.user.username))

											elif	'pvc_release'	in	str(request.POST):
															elevel=get_object_or_404(SupportedLevels,	pk=str(request.POST['pvc_release']).strip())
															elevel.sld=request.POST['sld']
															elevel.test_plan=request.POST['test_plan']
															elevel.management_node=request.POST['management_node']
															elevel.hmc_level=request.POST['hmc_level']
															elevel.vios_level=request.POST['vios_level']
															elevel.novalink_level=request.POST['novalink_level']
															elevel.novalink_os=request.POST['novalink_os']
															elevel.guest_os=request.POST['guest_os']
															elevel.storages=request.POST['storages']
															elevel.comments=request.POST['comments']
															elevel.save()
															logger.info("Edit Release {} is Successful, {}".format(str(request.POST['pvc_release']),request.user.username))

															return render(request,'hwt/ListSupportedLevels.html',{'supp_level':reversed(SupportedLevels.objects.all())})
				return render(request,'hwt/EditSupportedLevels.html',{'error':error})


@login_required
def	delete_supported_level(request):
				error=""
				if	request.POST:
											if	'drelease_id'	in	str(request.POST):
															try:
																							re=SupportedLevels.objects.get(pk=str(request.POST['drelease_id']).strip())
																							return render(request,'hwt/DeleteSupportedLevels.html',{'drelease':re})
															except	SupportedLevels.DoesNotExist:
																					error="Given	Release	Doesn't	exist"	
																					logger.error("Delete Release {} is Failed as the Release doesn't exists, {}".format(str(request.POST['drelease_id']),request.user.username))

											elif	'pvc_release'	in	str(request.POST):
															elevel=get_object_or_404(SupportedLevels,	pk=str(request.POST['pvc_release']).strip())
															elevel.delete()
															logger.info("Delete Release {} is Successful, {}".format(str(request.POST['pvc_release']),request.user.username))

															return render(request,'hwt/ListSupportedLevels.html',{'supp_level':reversed(SupportedLevels.objects.all())})
				return render(request,'hwt/DeleteSupportedLevels.html',{'error':error})


@login_required
def	list_supported_levels(request):
							return render(request,'hwt/ListSupportedLevels.html',{'supp_level':reversed(SupportedLevels.objects.all())})
							
@login_required
def	detail_list_supported_levels(request,	pvc_release):
							if	pvc_release:
												rel=get_object_or_404(SupportedLevels,	pk=str(pvc_release).strip())
												return	render(request,	'hwt/Detail_ListSupportedLevels.html',{'rel':rel})
							return render(request,'hwt/ListSupportedLevels.html',{'supp_level':reversed(SupportedLevels.objects.all())})

@login_required
def	list_hosts(request):
							#import pdb; pdb.set_trace()
							data=[]
							sq=[]
							hg=list(HostGroups.objects.values_list('name',flat=True))
							sq=list(Squads.objects.values_list('name',flat=True))
							for	obj	in	HostGroups.objects.all():
											dat1={}
											dat1['name']=obj.name
											sqd=[]
											htypes=[]
											htypes=HostTypes.objects.filter(host_group=obj.name).values_list('name',flat=True)
											#import pdb; pdb.set_trace()
											for	obj1	in	Squads.objects.all():
																sqd.append(len(Hosts.objects.filter(Q(cec_squad=obj1.name)	&	Q(cec_type__in=htypes))))
											dat1['data']=sqd
											data.append(dat1)
							return render(request,'hwt/ListAllHosts.html',{'hosts':Hosts.objects.all().order_by('cec_squad','cec_type'),'data':data,'squads':sq})

@login_required
def	list_hosts_type(request):
							hlist={}
							htypes=[]
							flist=[]
							data=[]
							for	obj	in	HostGroups.objects.all():
											hlist={}
											dat1=[]
											hlist['name']=obj.name
											dat1.append(obj.name)
											htypes=	HostTypes.objects.filter(host_group=obj.name).values_list('name',flat=True)
											hlist['host']=Hosts.objects.filter(cec_type__in=htypes).order_by('cec_squad','cec_type')
											dat1.append(len(hlist['host']))
											flist.append(hlist)
											data.append(dat1)
												
							#hgroups	=	HostGroups.objects.values_list('name',flat=True)
							#nlist=	zip(tuple(reversed(hlist)),tuple(reversed(htypes)))
							#p7s=Hosts.objects.filter(cec_type	=	'P7')
							#p8s=Hosts.objects.filter(cec_type	=	'P8')
							#p9s=Hosts.objects.filter(Q(cec_type='P9')|Q(cec_type='P9-Zeppelin')|Q(cec_type='P9-ZEPPELIN')|Q(cec_type='P9-ZZ')|Q(cec_type='P9-FLEETWOOD'))
							#kvms=Hosts.objects.filter(Q(cec_type	='KVM')|Q(cec_type='KVM-OPEN-POWER')|Q(cec_type='P9-BOSTON'))
							#return render(request,'hwt/ListHostsByType.html',{'p7s':p7s,'p8s':p8s,'p9s':p9s,'kvms':kvms})
							return render(request,'hwt/ListHostsByType.html',{'flist':flist,'data':data})

@login_required
def	list_hmcs(request):
							return render(request,'hwt/ListHMC.html',{'hmcs':Hmcs.objects.all()})

@login_required
def	list_storages(request):
							return render(request,'hwt/ListStorages.html',{'storages':Storages.objects.all()})

@login_required
def	list_fabrics(request):
							return render(request,'hwt/ListFabrics.html',{'fabrics':Fabrics.objects.all()})

@login_required
def	list_squad_wise(request):
								#infras=Hosts.objects.filter(cec_squad='Infra')
								#storages=Hosts.objects.filter(cec_squad='Storage')
								#frameworks=Hosts.objects.filter(cec_squad='Framework')
								#interfaces=Hosts.objects.filter(cec_squad='Interface')
								#techms=Hosts.objects.filter(cec_squad='Techm')
								#scales=Hosts.objects.filter(cec_squad='Scale')
								#devops=Hosts.objects.filter(cec_squad='Devops')
								#pools=Hosts.objects.filter(cec_squad='Pool')
								#unpings=Hosts.objects.filter(cec_squad='Unping')
								#context={'infras':infras,'storages':storages,'frameworks':frameworks,'interfaces':interfaces,'techms':techms,'scales':scales,'devops':devops,'pools':pools,'unpings':unpings}
								hcat=[]
								data=[]
								z=0
								hcat=HostGroups.objects.values_list('name',flat=True)
								squads=[]
								squads=list(Squads.objects.values_list('name',flat=True))
								for	obj	in	Squads.objects.all():
												drill={}
												htypes=[]
												hcount=[]
												hosts=[]
												dat1={}
												hosts=Hosts.objects.filter(cec_squad=obj.name)
												colors=['#7cb5ec','#90ed7d','#f1a65b','#8085e9','#f15c80','#e4d354','#2b908f','#f45b5b','#91e8e1','#434348']
												for	j	in	hcat:
																htypes=	HostTypes.objects.filter(host_group=j)
																nhosts=Hosts.objects.filter(Q(cec_squad=obj.name)	&	Q(cec_type__in=htypes))
																hcount.append(len(nhosts))
												drill['name']=obj.name
												drill['categories']=list(hcat)
												drill['data']=hcount
												dat1['y']=len(hosts)
												dat1['color']=colors[z]
												dat1['drilldown']=drill
												data.append(dat1)
												z+=1																				
								sqs=[]
								for	obj	in	Squads.objects.all():
											sq={}
											sq['name']=obj.name
											hosts=Hosts.objects.filter(cec_squad=obj.name).order_by('cec_type')
											sq['hosts']=hosts
											sqs.append(sq)
								return render(request,'hwt/ListSquadWiseAssignment.html',{'data':data,'squads':squads,'sqs':sqs})

@login_required
def	list_networks(request):
							return render(request,'hwt/ListNetworks.html',{'networks':Networks.objects.all(),'network_nodes':NetworkNodes.objects.all()})


@login_required
def	list_images(request):
							return render(request,'hwt/ListImages.html',{'images':Pvc_Images.objects.all()})





@login_required
def	snapshot(request):
				success=""
				error=""
				comment=""
				if	request.POST:
										if	'comments'	in	str(request.POST):
														comment=request.POST.get('comments')
										ns=datetime.datetime.now()
										sn=ns.timestamp()
										fpath="/root/proj/hw/dump/PVC"+str(sn)+".json"
										#fpath="PVC"+str(sn)+".json"
										
										subprocess.call(['source /root/proj/env/bin/activate'],shell=True)
										
										subprocess.call(['cd	/root/proj/hw'],shell=True)
										
										x=1
										x=subprocess.call(['python manage.py dumpdata --natural-foreign --natural-primary -e contenttypes -e auth.Permission --indent 2 -o	'+fpath], shell=True)
										if	not	x:
														snap=Snapshots(timestamp=ns,filepath=fpath,	comments=comment)
														snap.save()
														success="Successfully	Taken	Snapshot"
														logger.info(" Snapshot is successfully taken  at {}	,	{}".format(ns,request.user.username))

																	
										else:
													error="Encountered	issue	with	the	creation	of	Dump"
													logger.error(" Snapshot is Failed as there is Issue Encountered	with the creation of Dump	,	{}".format(request.user.username))

				return render(request,'hwt/SnapshotRestore.html',{'snapshots':reversed(Snapshots.objects.all()),'success':success,'error':error})
														
														
@login_required
def	restore(request):
				success=""
				error=""
				#import pdb; pdb.set_trace()
				if	request.POST:
															if	'optradio'	in	str(request.POST):
																			fe=str(request.POST['optradio'])
																			if	os.path.isfile(fe):
																									subprocess.call('cd	/root/hw',shell=True)
																									if	not	subprocess.call('python manage.py loaddata	'+fe,shell=True):
																												success="Successfully	Restored"
																												logger.info(" Restore is Successfully done with {},	{}".format(fe,request.user.username))

																			else:	
																									error="File	Doesn't	exist"
																									logger.error(" Restore  Failed as the File doesn't exist{},	{}".format(fe,request.user.username))

															else:
																			error="You	have	NOT	selected	any	snap	shot	to	restore"
																			logger.error(" Restore  Failed as no File is selected	to	restore,	{}".format(request.user.username))

				return render(request,'hwt/SnapshotRestore.html',{'snapshots':reversed(Snapshots.objects.all()),'success':success,'error':error})

	
	
@login_required
def	email_notify(request):
				success=""
				error=""
				if	request.POST:
							subject = request.POST['subject']
							message = request.POST['message']
							to_email=	request.POST['to_email']
							if subject and message	and	to_email:
											try:
															send_mail(subject, message, "",	[to_email],	fail_silently=False )
															success="Email	is	successfully	sent"
															logger.info(" Email is Successfully Sent to {} ,	{}".format(str(to_email),request.user.username))

											except :
																error="Not	able	to	send	email"
																logger.error(" Send Email Failed ,	{}".format(request.user.username))

							else:
									error="Make	sure	all	the	fields	are	given	valid	values"
									logger.error(" Send Email Failed- Make	sure	all	the	fields	are	given	valid	values ,	{}".format(request.user.username))

				return render(request,'hwt/EmailNotify.html',{'success':success,'error':error})


@login_required
def	host_groups(request):
				success=""
				error=""
				if	request.POST:
								if	str(request.POST['hg_name']).strip():
														if	bool(re.search(r"\s", str(request.POST['hg_name']).strip())) :
																		error="Name	should	be	single	word	without	any	spaces	in	between"
																		logger.error(" Add Host Group failed as the Name	should	be	single	word	without	any	spaces	in	between {},	{}".format(str(request.POST['hg_name']),request.user.username))

														else:
																		hgname=str(request.POST['hg_name']).strip()
																		hgdesc=request.POST['hg_descreption']
																		hg=HostGroups(name=hgname,descreption=hgdesc)
																		hg.save()
																		success="Host	group	is	successfully	created"
																		logger.info(" Add Host Group {} is Successful,	{}".format(str(request.POST['hg_name']),request.user.username))

								else:
												error="Mandatory	fields	are	missing"
				return render(request,'hwt/HostGroupsTypes.html',{'hgs':HostGroups.objects.all(),'hts':HostTypes.objects.all(),	'hgsuccess':success,	'hgerror':error})

@login_required
def	delete_host_groups(request):
				success=""
				error=""
				if	'hgopt'	in	request.POST:
							try:
											dg	=	HostGroups.objects.get(pk=request.POST['hgopt'])
											dg.delete()
											success="Successfully	deleted	Host	Group"
											logger.info(" Delete Host Group {} is Successful,	{}".format(str(request.POST['hgopt']),request.user.username))

							except	HostGroups.DoesNotExist:	
												error="Host	Group	doesn't	exist"
												logger.error(" Delete Host Group {} is failed as the Host group doesn't exist,	{}".format(str(request.POST['hgopt']),request.user.username))

				else:
									error="Host	group	is	not	selected"
									logger.error(" Delete Host Group  is failed as the Host group is not selected,	{}".format(request.user.username))

				return render(request,'hwt/HostGroupsTypes.html',{'hgs':HostGroups.objects.all(),'hts':HostTypes.objects.all(),	'hgsuccess':success,	'hgerror':error})

@login_required
def	host_types(request):
				success=""
				error=""
				if	request.POST:
								if	'ht_name'	in	request.POST	and	'htgroup'	in	request.POST:
												if	str(request.POST['ht_name']).strip()	and	str(request.POST['htgroup']).strip():
															htname=str(request.POST['ht_name']).strip()
															htdesc=request.POST['ht_descreption']
															htgroup=get_object_or_404(HostGroups,	pk=str(request.POST['htgroup']).strip())
															ht=HostTypes(name=htname,descreption=htdesc,host_group=htgroup)
															ht.save()
															logger.info(" Add Host Type {} is Successful,	{}".format(str(request.POST['ht_name']),request.user.username))
															success="Host	Type	is	successfully	created"
												else:
															error="Mandatory	fields	are	missing"
				return render(request,'hwt/HostGroupsTypes.html',{'hgs':HostGroups.objects.all(),'hts':HostTypes.objects.all(),	'htsuccess':success,	'hterror':error})


@login_required
def	delete_host_types(request):
				success=""
				error=""
				if	'htopt'	in	request.POST:
							try:
											dt	=	HostTypes.objects.get(pk=request.POST['htopt'])
											dt.delete()
											success="Successfully	deleted	Host	Type"
											logger.info(" Delete Host Type {} is Successful,	{}".format(str(request.POST['htopt']),request.user.username))

							except	HostTypes.DoesNotExist:	
												error="Host	Type	doesn't	exist"
												logger.error(" Delete Host Type is Failed as the Host Type doesn't exist,	{}".format(request.user.username))

				else:
									error="Host	Type	is	not	selected"
									logger.error(" Delete Host Type is Failed as the Host Type is not selected,	{}".format(request.user.username))

				return render(request,'hwt/HostGroupsTypes.html',{'hgs':HostGroups.objects.all(),'hts':HostTypes.objects.all(),	'htsuccess':success,	'hterror':error})

@login_required
def list_scheduler(request):
		success=""
		error=""
		ipadd=""
		stat=""
		notify_list=""
		if request.POST:
				if str(request.POST['ip_add']).strip():
								if	bool(re.search(r"\s", str(request.POST['ip_add']).strip())) :
										error="IP address should not have any spaces	in	between"
										logger.error(" Add Scheduled notification for {} is Failed as IP address contains space in between,	{}".format(str(request.POST['ip_add']),request.user.username))
								else:
										ipadd=request.POST['ip_add']
										stat=request.POST['status']
										notify_list=request.POST['notify_ids']
										sq=Schedules(ip_add=ipadd,timestamp=datetime.datetime.now(),status=stat,notify_date=datetime.date.today(),notified_today=False,notify_id=notify_list)
										sq.save()
										logger.info(" Add Schduled Notification for  {} is Successful,	{}".format(str(request.POST['ip_add']),request.user.username))
										success="Schedule Notification	is	successfully	created"
				else:
										error="Mandatory	fields	are	missing"
															
		return render(request,'hwt/list_scheduler.html',{'schedules':Schedules.objects.all(),'success':success,'error':error})
		 
@login_required
def delete_scheduler(request):
		success=""
		error=""
		if	'scopt'	in	request.POST:
							try:
											sq	=	Schedules.objects.get(pk=request.POST['scopt'])
											sq.delete()
											success="Successfully	deleted	Schedule"
											logger.info(" Delete Schedule {} is Successful,	{}".format(str(request.POST['scopt']),request.user.username))

							except	Squads.DoesNotExist:	
												error="Squad	doesn't	exist"
												logger.error(" Delete Squad {} is Failed as Squad	doesn't	exist,	{}".format(str(request.POST['scopt']),request.user.username))

		else:
									error="Schedule	is	not	selected"
									logger.error(" Delete Schedule is Failed as Schedule list item is not selected,	{}".format(request.user.username))

		return render(request,'hwt/list_scheduler.html',{'schedules':Schedules.objects.all(),'success':success,'error':error})


@login_required
def	squad_update(request):
				success=""
				error=""
				if	request.POST:
								if	str(request.POST['sq_name']).strip():
															if	bool(re.search(r"\s", str(request.POST['sq_name']).strip())) :
																		error="Name	should	be	single	word	without	any	spaces	in	between"
																		logger.error(" Add Squad {} is Failed as Name	should	be	single	word	without	any	spaces	in	between,	{}".format(str(request.POST['sq_name']),request.user.username))
															else:
																			sqname=request.POST['sq_name']
																			sqdesc=request.POST['sq_descreption']
																			sq=Squads(name=sqname,descreption=sqdesc)
																			sq.save()
																			logger.info(" Add Squad {} is Successful,	{}".format(str(request.POST['sq_name']),request.user.username))

																			success="Squad	is	successfully	created"
								else:
												error="Mandatory	fields	are	missing"
				return render(request,'hwt/SquadData.html',{'sqs':Squads.objects.all(),	'success':success,	'error':error})

@login_required
def	squad_delete(request):
				success=""
				error=""
				if	'sqopt'	in	request.POST:
							try:
											sq	=	Squads.objects.get(pk=request.POST['sqopt'])
											sq.delete()
											success="Successfully	deleted	Squad"
											logger.info(" Delete Squad {} is Successful,	{}".format(str(request.POST['sqopt']),request.user.username))

							except	Squads.DoesNotExist:	
												error="Squad	doesn't	exist"
												logger.error(" Delete Squad {} is Failed as Squad	doesn't	exist,	{}".format(str(request.POST['sqopt']),request.user.username))

				else:
									error="Squad	is	not	selected"
									logger.error(" Delete Squad is Failed as Squad is not selected,	{}".format(request.user.username))

				return render(request,'hwt/SquadData.html',{'sqs':Squads.objects.all(),	'success':success,	'error':error})

@login_required
def	import_data(request):
				success=""
				error=""
				exclusions=[]
				inclusions=[]
				htexclusions=[]
				sqexclusions=[]
				Templates={
				'Hosts':['FSP_/_BMC_IP','Serial','Location','CEC_Name','FSP_Credentials',	'Host_Type',	'Firmware_Level',	'HMC_IP',	'Novalink_IP',	'VIOS1',	'VIOS2',	'Proc_/_Memory',	'Model',	'Network_Ports',	'Fabric_Switch1',	'Fabric_Switch2',	'Primary_VLAN',	'Connected_Storages',	'Owner',	'Squad',	'LAB'],
				'HMCs':['HMC_Public_IP','Serial','Location','HMC_Credentials','HMC_Private_IP','Public_/_Private','IMM_IP','IMM_Credentials','MTMS','Owner','Squad','LAB'],
				'Storages':['Storage_IP','Serial','Location','Storage_Type','Storage_Credentials','Storage_Port','Storage_REST_API_Server','Storage_LDEV_ID_range','Capacity_in_TB','Service_IP','LAB'],
				'Fabrics':['Fabric_IP','Serial','Location','Credentials','Brocade_/_CISCO','Primary_/_Secondary','Fabric_Port','VSAN_ID','LAB','Connected_Storwises','Connected_DS8K','Connected_EMC_VMAX','Connected_EMC_VNX','Connected_Hitachi','Connected_XIV']
				}
				if	request.POST:
																				rtype=request.POST['rtype']
																				rfile1=request.FILES['excel_file']
																				if	not	rfile1.name.endswith('xlsx'):
																							error="Pls	upload	the	xlsx	file	only	"
																							logger.error(" Import {} data is Failed -Pls upload	the	xlsx	file	only,	{}".format(str(request.POST['rtype']),request.user.username))

																				else:
																							exclusions=[]
																							inclusions=[]
																							htexclusions=[]
																							sqexclusions=[]
																							wb = openpyxl.load_workbook(rfile1)
																											
																							if	rtype	==	"Hosts":
																											if	'Hosts'	not	in	wb.sheetnames:
																															error=	"There	is	no	Sheet	with	the	name	Hosts"
																															logger.error(" Import {} data is Failed- There	is	no	Sheet	with	the	name	Hosts,	{}".format(str(request.POST['rtype']),request.user.username))

																											else:	
																														ws	=	wb['Hosts']
																														collist=Templates[rtype]
																														first_row = []
																														flag=0
																														i=0
																														while	i	<	len(collist):
																																		if	collist[i]!=str(ws.cell(row=1,column=(i+1)).value).strip():
																																					flag=1
																																					break
																																		i+=1
																														if	flag==0:
																																		htypes=[]
																																		htypes=HostTypes.objects.values_list('name',flat=True)
																																		squads=[]
																																		squads=Squads.objects.values_list('name',flat=True)
																																		mxrows = (ws.max_row)-1
																																		count=0
																																		for	row	in ws.iter_rows(min_row=2):
																																						if	row[0].value	is	not	None:
																																										row_data=[]
																																										ht=""
																																										sq=""
																																										for	cell	in	row:
																																														x=str(cell.value).strip()
																																														if	cell.value	is	None:
																																																			x=""																																																		
																																														row_data.append(x)
																																										if	str(row_data[0])	!=	"":
																																														try:
																																																				obj = Hosts.objects.get(pk=row_data[0])
																																																				logger.error("Import data of type {} with IP {} is Failed as the resource already exists,	{}".format(str(request.POST['rtype']),str(row_data[0]),request.user.username))
																																																				exclusions.append(row_data[0])
																																														except	Hosts.DoesNotExist:
																																																					ht=str(row_data[5]).strip()
																																																					sq=str(row_data[19]).strip()
																																																					if	not	sq:
																																																								sq="Pool"
																																																					if	ht	==	""	or	ht	not	in	htypes:
																																																								logger.error("Import data of type {} with IP {} is Failed as Host Type doesn't exist,	{}".format(str(request.POST['rtype']),str(row_data[0]),request.user.username))
																																																								htexclusions.append({'ip':row_data[0],'htype':ht})
																																																					elif	sq	not	in	squads:
																																																									logger.error("Import data of type {} with IP {} is Failed as Squad doesn't exist,	{}".format(str(request.POST['rtype']),str(row_data[0]),request.user.username))
																																																									sqexclusions.append({'ip':row_data[0],'squad':sq})
																																																					else:
																																																									obj=Hosts(
																																																									cec_fsp_bmc_ip=row_data[0].strip(),
																																																									cec_ip_ping=pingStatus(row_data[0].strip()),
																																																									cec_serial=row_data[1],
																																																									cec_location=row_data[2],
																																																									cec_name=row_data[3],
																																																									cec_fsp_credentials=row_data[4],
																																																									cec_type=ht,
																																																									cec_firmware=row_data[6],
																																																									cec_hmc_ip=row_data[7],
																																																									cec_neo_ip=row_data[8],
																																																									cec_vios1=row_data[9],
																																																									cec_vios2=row_data[10],
																																																									cec_proc_memory=row_data[11],
																																																									cec_model=row_data[12],
																																																									cec_network_ports=row_data[13],
																																																									cec_fabric1=row_data[14],
																																																									cec_fabric2=row_data[15],
																																																									cec_pvlan=row_data[16],
																																																									cec_storages=row_data[17],
																																																									cec_owner=row_data[18],
																																																									cec_squad=sq,
																																																									cec_lab=row_data[20])
																																																									obj.save()
																																																									logger.info("Import data of type {} with IP {} is Done ,	{}".format(str(request.POST['rtype']),str(row_data[0]),request.user.username))
																																																									inclusions.append(row_data[0])																								
																																						#count+=1
																																						#job = updateTask.delay(int((count / mxrows)*100))
																																						##task_status(job.id)
																																						#return render(request,'hwt/ImportData.html',{'task_id':job.id})													

																														else:
																																	error=	"Uploaded	file	is	not	having	right	columns"
																																	logger.error(" Import {} data is Failed -Uploaded	file	is	not	having	right	columns,	{}".format(str(request.POST['rtype']),request.user.username))

																																	
################################################################################################################################																																	
																							elif	rtype	==	"HMCs":
																											if	'HMCs'	not	in	wb.sheetnames:
																															error=	"There	is	no	Sheet	with	the	name	HMCs"
																															logger.error(" Import {} data is Failed- There	is	no	Sheet	with	the	name	HMCs,	{}".format(str(request.POST['rtype']),request.user.username))

																											else:	
																														ws	=	wb['HMCs']
																														collist=Templates[rtype]
																														flag=0
																														i=0
																														while	i	<	len(collist):
																																		if	collist[i]!=str(ws.cell(row=1,column=(i+1)).value).strip():
																																					flag=1
																																					break
																																		i+=1
																														if	flag==0:
																																		for	row	in	ws.iter_rows(min_row=2):
																																						if	row[0].value	is	not	None:
																																										row_data=[]
																																										for	cell	in	row:
																																															row_data.append(str(cell.value))
																																										if	(str(row_data[0])	!=	""):
																																											try:
																																													obj = Hmcs.objects.get(pk=str(row_data[0]).strip())
																																													exclusions.append(row_data[0])
																																													logger.error("Import data of type {} with IP {} is Failed as the resource already exists,	{}".format(str(request.POST['rtype']),row_data[0],request.user.username))						

																																											except Hmcs.DoesNotExist:						
																																													obj=Hmcs(
																																																hmc_public_ip=str(row_data[0]).strip(),
																																																					hmc_ip_ping=pingStatus(str(row_data[0]).strip()),
																																																					hmc_serial=row_data[1],
																																																					hmc_location=row_data[2],
																																																					hmc_credentials=row_data[3],
																																																					hmc_private_ip=row_data[4],
																																																					hmc_public_private=row_data[5],
																																																					hmc_imm_ip=row_data[6],
																																																					hmc_imm_credentials=row_data[7],
																																																					hmc_mtms=row_data[8],
																																																					hmc_owner=row_data[9],
																																																					hmc_squad=row_data[10],
																																																					hmclab=row_data[11])
																																													obj.save()
																																													logger.info("Import data of type {} with IP {} is Done ,	{}".format(str(request.POST['rtype']),str(row_data[0]),request.user.username))
																																													inclusions.append(row_data[0])																								
																														else:
																																	error=	"Uploaded	file	is	not	having	right	columns"
																																	logger.error(" Import {} data is Failed -Uploaded	file	is	not	having	right	columns,	{}".format(str(request.POST['rtype']),request.user.username))

################################################################################################################################																																	
																							elif	rtype	==	"Storages":
																											if	'Storages'	not	in	wb.sheetnames:
																															error=	"There	is	no	Sheet	with	the	name	Storages"
																															logger.error(" Import {} data is Failed- There	is	no	Sheet	with	the	name	Storages,	{}".format(str(request.POST['rtype']),request.user.username))

																											else:	
																														ws	=	wb['Storages']
																														collist=Templates[rtype]
																														first_row = []
																														flag=0
																														i=0
																														while	i	<	len(collist):
																																		if	collist[i]!=str(ws.cell(row=1,column=(i+1)).value).strip():
																																					flag=1
																																					break
																																		i+=1				
																														if	flag==0:
																																		for	row	in	ws.iter_rows(min_row=2):
																																						if	row[0].value	is	not	None:
																																										row_data=[]
																																										for	cell	in	row:
																																															row_data.append(str(cell.value))
																																										if	str(row_data[0])	!=	"":
																																														try:
																																																	obj = Storages.objects.get(pk=str(row_data[0]).strip())
																																																	logger.error(" Import data of type {} with IP {} is Failed as the resource already exists,	{}".format(str(request.POST['rtype']),str(row_data[0]),request.user.username))
																																																	exclusions.append(row_data[0])
																																														except Storages.DoesNotExist:
																																																					obj=Storages(
																																																					storage_ip=str(row_data[0]).strip(),
																																																					storage_ip_ping=pingStatus(str(row_data[0]).strip()),
																																																					storage_serial=row_data[1],
																																																					storage_location=row_data[2],
																																																					storage_type=row_data[3],
																																																					storage_credentials=row_data[4],
																																																					storage_port=row_data[5],
																																																					storage_rest_api_server=row_data[6],
																																																					storage_ldev_id_range=row_data[7],
																																																					storage_capacity=row_data[8],
																																																					storage_service_ip=row_data[9],
																																																					storage_lab=row_data[10])
																																																					obj.save()
																																																					logger.info("Import data of type {} with IP {} is Done ,	{}".format(str(request.POST['rtype']),str(row_data[0]),request.user.username))
																																																					inclusions.append(row_data[0])																								
																														else:
																																	error=	"Uploaded	file	is	not	having	right	columns"
																																	logger.error(" Import {} data is Failed -Uploaded	file	is	not	having	right	columns,	{}".format(str(request.POST['rtype']),request.user.username))

################################################################################################################################
																							elif	rtype	==	"Fabrics":
																											if	'Fabrics'	not	in	wb.sheetnames:
																															error=	"There	is	no	Sheet	with	the	name	Fabrics"
																															logger.error("Import data of type {} Failed as - There	is	no	Sheet	with	the	name	Fabrics,	{}".format(str(request.POST['rtype']),request.user.username))

																											else:		
																														ws	=	wb['Fabrics']
																														collist=Templates[rtype]
																														first_row = []
																														flag=0
																														i=0
																														while	i	<	len(collist):
																																		if	collist[i]!=str(ws.cell(row=1,column=(i+1)).value).strip():
																																					flag=1
																																					break
																																		i+=1				
																														if	flag==0:
																																		for	row	in	ws.iter_rows(min_row=2):
																																						if	row[0].value	is	not	None:
																																										row_data=[]
																																										for	cell	in	row:
																																															row_data.append(str(cell.value))
																																										if	str(row_data[0])	!=	"":
																																														try:
																																																	obj = Fabrics.objects.get(pk=str(row_data[0]).strip())
																																																	logger.error(" Import data of type {} with IP {} is Failed as the resource already exists,	{}".format(str(request.POST['rtype']),str(row_data[0]),request.user.username))
																																																	exclusions.append(row_data[0])
																																																	
																																														except Fabrics.DoesNotExist:
																																																					obj=Fabrics(
																																																					fabric_ip=str(row_data[0]).strip(),
																																																					fabric_ip_ping=pingStatus(str(row_data[0]).strip()),
																																																					fabric_serial=row_data[1],
																																																					fabric_location=row_data[2],
																																																					fabric_credentials=row_data[3],
																																																					fabric_type=row_data[4],
																																																					fabric_primary=row_data[5],
																																																					fabric_primary_port=row_data[6],
																																																					fabric_vsan_id=row_data[7],
																																																					fabric_lab=row_data[8],
																																																					fabric_storwise=row_data[9],
																																																					fabric_ds8k=row_data[10],
																																																					fabric_emc_vmax=row_data[11],
																																																					fabric_emc_vnx=row_data[12],
																																																					fabric_hitachi=row_data[13],
																																																					fabric_xiv=row_data[14])
																																																					obj.save()
																																																					logger.info("Import data of type {} with IP {} is Done ,	{}".format(str(request.POST['rtype']),str(row_data[0]),request.user.username))
																																																					inclusions.append(row_data[0])																								
																														else:
																																	error=	"Uploaded	file	is	not	having	right	columns"
																																	logger.error(" Import {} data is Failed -Uploaded	file	is	not	having	right	columns,	{}".format(str(request.POST['rtype']),request.user.username))

																							
				return render(request,'hwt/ImportData.html',{'success':success,	'error':error,'inclusions':inclusions,	'exclusions':exclusions,'htexclusions':htexclusions,'sqexclusions':sqexclusions})

@login_required
def	list_logs(request):
						error=""
						fe='/root/proj/hw/LOG/debug.log'
						if	os.path.isfile(fe):
									f=open(fe,'r')
									logs=[]
									log={}
									#import pdb; pdb.set_trace()
									myList=['INFO','ERROR']
									for	x	in	f:
														#y=f.readline()
												if any(y in str(x) for y in myList):
														log={}
														tmp=[]
														tmp=x.split(',')
														if	len(tmp)==5:
																if not "reloading" in str(tmp[4]):
																		log['level']=tmp[0]
																		log['time']=tmp[1]
																		log['message']=tmp[3]
																		log['user']=tmp[4]
																		logs.append(log)
									f.close()
						return	render(request,'hwt/Logs.html',{'logs':reversed(logs),	'error':error})

				
@login_required
def	task_status(request,task_id):	
	result=AsyncResult(task_id)
	response_data={
		'state':result.state,
		'progress':result.info['current'],

	}
	#return HttpResponse(json.dumps(response_data),content_type='application/json')
	return render(request,'hwt/ImportData.html',{})


